import openpyxl
from io import BytesIO
from django.http import HttpResponse
from datetime import datetime
from rest_framework.decorators import action
from rest_framework.response import Response
from openpyxl.utils import get_column_letter


class ExportExcelMixin:
    @action(detail=True, methods=['get'], url_path='exportar_excel')
    def exportar_excel(self, request, pk=None):
        instance = self.get_object()
        modelo_nombre = instance._meta.model_name
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte"
        
        fields = instance._meta.get_fields()
        col_num = 1
        headers = []
        for field in fields:
            if not field.many_to_one and not field.many_to_many and not field.one_to_many:
                headers.append(field.name)
                ws.cell(row=1, column=col_num, value=field.verbose_name.title())
                col_num += 1
        
        col_num = 1
        for header in headers:
            value = getattr(instance, header, '')
            if isinstance(value, datetime):
                value = value.strftime('%d/%m/%Y %H:%M')
            ws.cell(row=2, column=col_num, value=value)
            col_num += 1
            
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{modelo_nombre}_{instance.pk}.xlsx"'
        return response


class CustomResponseMixin:
    
    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            paginated = self.paginator.get_paginated_response(serializer.data).data
            return Response({
                'success': True,
                'status': 200,
                'message': 'Petición exitosa',
                'count': paginated.get('count'),
                'next': paginated.get('next'),
                'previous': paginated.get('previous'),
                'data': paginated.get('results'),
            })
        serializer = self.get_serializer(queryset, many=True)
        return Response({
            'success': True,
            'status': 200,
            'message': 'Petición exitosa',
            'data': serializer.data
        })

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return Response({
            'success': True,
            'status': 200,
            'message': 'Petición exitosa',
            'data': serializer.data
        })

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({
                'success': True,
                'status': 201,
                'message': 'Registro creado exitosamente',
                'data': serializer.data
            }, status=201)
        return Response({
            'success': False,
            'status': 400,
            'message': 'Error de validación',
            'errors': serializer.errors
        }, status=400)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        if serializer.is_valid():
            serializer.save()
            return Response({
                'success': True,
                'status': 200,
                'message': 'Registro actualizado exitosamente',
                'data': serializer.data
            })
        return Response({
            'success': False,
            'status': 400,
            'message': 'Error de validación',
            'errors': serializer.errors
        }, status=400)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.delete()
        return Response({
            'success': True,
            'status': 200,
            'message': 'Registro eliminado exitosamente'
        })