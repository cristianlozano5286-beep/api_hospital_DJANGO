from rest_framework import viewsets, filters, status
from rest_framework.decorators import action
from rest_framework.response import Response
from drf_spectacular.utils import extend_schema, extend_schema_view
from .models import Paciente
from .serializers import PacienteSerializer
from rest_framework.renderers import BaseRenderer
from rest_framework.permissions import AllowAny
from rest_framework.decorators import permission_classes
from io import BytesIO
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime


class ExcelRenderer(BaseRenderer):
    media_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    format = 'xlsx'
    charset = None

    def render(self, data, accepted_media_type=None, renderer_context=None):
        return data


@extend_schema_view(
    list=extend_schema(summary='Listar pacientes', tags=['Pacientes']),
    create=extend_schema(summary='Registrar paciente', tags=['Pacientes']),
    retrieve=extend_schema(summary='Obtener paciente', tags=['Pacientes']),
    update=extend_schema(summary='Actualizar paciente', tags=['Pacientes']),
    partial_update=extend_schema(summary='Actualizar parcialmente paciente', tags=['Pacientes']),
    destroy=extend_schema(summary='Desactivar paciente', tags=['Pacientes']),
    exportar_excel=extend_schema(summary="Exportar a Excel", tags=['Pacientes']),
)
class PacienteViewSet(viewsets.ModelViewSet):
    serializer_class = PacienteSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['nombres', 'apellidos', 'documento_identidad', 'correo_electronico', 'eps']
    ordering_fields = ['apellidos', 'nombres', 'fecha_nacimiento', 'fecha_creacion']
    ordering = ['apellidos']

    def get_queryset(self):
        qs = Paciente.objects.all()
        activo = self.request.query_params.get('activo')
        if activo is not None:
            qs = qs.filter(activo=activo.lower() == 'true')
        return qs

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.desactivar()
        return Response(
            {'mensaje': f'Paciente "{instance.nombre_completo}" desactivado.'},
            status=status.HTTP_200_OK,
        )

    @extend_schema(summary='Activar paciente', tags=['Pacientes'])
    @action(detail=True, methods=['patch'], url_path='activar')
    def activar(self, request, pk=None):
        instance = self.get_object()
        instance.activar()
        return Response(PacienteSerializer(instance).data)

    @extend_schema(summary='Historial de citas del paciente', tags=['Pacientes'])
    @action(detail=True, methods=['get'], url_path='citas')
    def citas(self, request, pk=None):
        from apps.citas.models import Cita
        from apps.citas.serializers import CitaSerializer
        paciente = self.get_object()
        citas = Cita.objects.filter(paciente=paciente).select_related('medico', 'medico__especialidad')
        return Response(CitaSerializer(citas, many=True).data)
    
    @extend_schema(
        summary="Exportar perfil de paciente a Excel",
        tags=['Pacientes'],
        responses={200: "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
    )
    @action(detail=True, methods=['get'], renderer_classes=[ExcelRenderer])
    @permission_classes([AllowAny])
    def exportar_excel(self, request, pk=None):
        paciente = self.get_object()
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Paciente"
        
        # Obtener todos los campos del modelo
        fields = paciente._meta.get_fields()
        
        # Crear encabezados
        col_num = 1
        headers = []
        for field in fields:
            if not field.many_to_one and not field.many_to_many and not field.one_to_many:
                headers.append(field.name)
                ws.cell(row=1, column=col_num, value=field.verbose_name.title())
                col_num += 1
        
        # Llenar datos
        col_num = 1
        for header in headers:
            value = getattr(paciente, header, '')
            if isinstance(value, datetime):
                value = value.strftime('%d/%m/%Y %H:%M')
            ws.cell(row=2, column=col_num, value=value)
            col_num += 1
        
        # Ajustar ancho de columnas
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Guardar en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return Response(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename="paciente_{paciente.documento_identidad}.xlsx"'}
        )
