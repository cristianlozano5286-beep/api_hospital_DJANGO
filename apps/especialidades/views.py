from django.template.loader import render_to_string
from django.http import HttpResponse
from weasyprint import HTML
from rest_framework import viewsets, filters, status
from rest_framework.decorators import action, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import AllowAny
from drf_spectacular.utils import extend_schema, extend_schema_view
from .models import Especialidad
from .serializers import EspecialidadSerializer
from io import BytesIO
import openpyxl
from datetime import datetime

@extend_schema_view(
    list=extend_schema(summary='Listar especialidades', tags=['Especialidades']),
    create=extend_schema(summary='Crear especialidad', tags=['Especialidades']),
    retrieve=extend_schema(summary='Obtener especialidad', tags=['Especialidades']),
    update=extend_schema(summary='Actualizar especialidad', tags=['Especialidades']),
    partial_update=extend_schema(summary='Actualizar parcialmente especialidad', tags=['Especialidades']),
    destroy=extend_schema(summary='Eliminar especialidad', tags=['Especialidades']),
    exportar_pdf=extend_schema(summary="Exportar a PDF", tags=['Especialidades']),
    exportar_excel=extend_schema(summary="Exportar a Excel", tags=['Especialidades']),
)
class EspecialidadViewSet(viewsets.ModelViewSet):
    serializer_class = EspecialidadSerializer
    queryset = Especialidad.objects.all()
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['nombre', 'codigo']
    ordering_fields = ['nombre', 'codigo', 'fecha_creacion']
    ordering = ['nombre']

    @action(detail=True, methods=['get'], url_path='exportar_pdf')
    @permission_classes([AllowAny])
    def exportar_pdf(self, request, pk=None):
        try:
            instance = self.get_object()
            html_string = render_to_string('especialidad/reporte.html', {'object': instance})
            pdf = HTML(string=html_string).write_pdf()
            
            response = HttpResponse(pdf, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="especialidad_{instance.id}.pdf"'
            return response
        except Exception as e:
            return HttpResponse(f"Error generando PDF: {str(e)}", status=500)

    @action(detail=True, methods=['get'], url_path='exportar_excel')
    @permission_classes([AllowAny])
    def exportar_excel(self, request, pk=None):
        try:
            instance = self.get_object()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Especialidad"
            
            fields = instance._meta.get_fields()
            col_num = 1
            headers = []
            for field in fields:
                if not field.is_relation:
                    headers.append(field.name)
                    ws.cell(row=1, column=col_num, value=field.verbose_name.title())
                    col_num += 1
            
            col_num = 1
            for header in headers:
                value = getattr(instance, header, '')
                if isinstance(value, datetime):
                    value = value.strftime('%d/%m/%Y %H:%M')
                ws.cell(row=2, column=col_num, value=value)
                col_num += 1
            
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                ws.column_dimensions[column].width = min(max_length + 2, 50)
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="especialidad_{instance.id}.xlsx"'
            return response
        except Exception as e:
            return HttpResponse(f"Error generando Excel: {str(e)}", status=500)

    def get_queryset(self):
        qs = Especialidad.objects.all()
        activo = self.request.query_params.get('activo')
        if activo is not None:
            qs = qs.filter(activo=activo.lower() == 'true')
        return qs

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.desactivar()
        return Response({'mensaje': f'Especialidad "{instance.nombre}" desactivada.'}, status=status.HTTP_200_OK)

    @extend_schema(summary='Activar especialidad', tags=['Especialidades'])
    @action(detail=True, methods=['patch'], url_path='activar')
    def activar(self, request, pk=None):
        instance = self.get_object()
        instance.activar()
        return Response(EspecialidadSerializer(instance).data)